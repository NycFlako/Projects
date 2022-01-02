from cmu_112_graphics import *
from tkinter import *
from PIL import Image
import random, math, copy, time, pygame
from datetime import date
from openpyxl import Workbook, load_workbook
from Components import *

Cmp = Components()

"""
Icons used were gathered from flaticon.com
"""
'''
Documentation for Workbooks on openpyxl: https://openpyxl.readthedocs.io/en/stable/tutorial.html


######################## Worksheet documentation ##############################
Create new workbook -> wb = Workbook()

Get first sheet -> ws = wb.active

To create a new sheet on workbook -> ws2 = wb.create_sheet(title = "Mysheet", index)
By default create_sheet will create it at the end if no index is specified.
Index 0 is the first position and index -1 is the penultimate

To change title of a sheet -> ws.title = "New Title"

Sheets can be obtained from workbook with their name as key -> ws3 = wb["New Title"]

To see list of all of the sheets in a workbook -> print(wb.sheetnames)

You can also loop through worksheets -> for sheet in wb:


########################## Cells documentation ################################

When a worksheet is created it contains no cells, they must be created first


Cells can be accessed as keys on the ws -> c = ws['A4']

Cells can also be accessed by cells using row and col -> d = ws.cell(row=4, column=2, value=10)

Ranges of cells can be obtained through slicing -> cell_range = ws['A1':'C2']
Example for columns: colC = ws['C'], col_range = ws['C:D']
Example for rows: row10 = ws[10], row_range = ws[5:10]

To iterate through all of the rows in a file -> ws.rows

We can assign values to cells directly -> c.value = 'hello, world'

##########################  Working with Workbooks #############################

To save the current Workbook -> wb.save('balances.xlsx') 
This will overwrite any existing workbook with the given name.

To load up a workbook -> wb2 = load_workbook('test.xlsx')

############################ Tricks and Tips #################################

Set date and time using Python date -> ws['A1'] = datetime.datetime(2010, 7, 21)

Adding a formula to a cell -> ws["A1"] = "=SUM(1, 1)"

'''
pygame.mixer.init()

########## Initializing the system ############# 
class InitializeSystem(Mode):
    def getDate(mode):
        today = date.today()
        day, dayNum = today.strftime("%A"), today.strftime("%d")
        monthFull, yearFull = today.strftime("%B"), today.strftime("%Y")
        workbook = monthFull+"-"+yearFull
        title = day+" "+dayNum
        return workbook, title

    def getLog(mode, path):
        (workBookPath, sheetTitle) = mode.getDate()
        try:
            workBook = load_workbook(path+workBookPath)
        except:
            workBook = Workbook()
            workBook.save(path+workBookPath+".xlsx")
        newSheet = workBook.create_sheet(title = sheetTitle)
        return (workBook, newSheet)

    def getMenuDrinks(mode):
        return mode.menu[mode.items[0]]

    def getMenuSandwich(mode):
        return mode.menu[mode.items[1]]

    def getMenuDessert(mode):
        return mode.menu[mode.items[2]]

    def getMenuOther(mode):
        return mode.menu[mode.items[3]]
        
    def parseMenuLine(mode, s):
        result, L = "", list()
        items = s.split(",")
        for i in range(len(items)):
            elem = items[i].strip()
            if(i == 0):
                result = elem
            else:
                L.append(elem)
        return result, L

    def getMenu(mode, path):
        file = open(path, "r")
        menu = dict()
        for line in file.readlines():
            if("," not in line):
                section = line.strip()
                newDict = dict()
                menu[section] = newDict
            else:
                kind, options = mode.parseMenuLine(line)
                if(kind in newDict):
                    print("Already Inside")
                else:
                    newDict[kind] = {"title": kind, "items": options}
        return menu

    def appStarted(mode):
        mode.items = ["Drinks", "Sandwich", "Dessert", "Other"]
        mode.menu = mode.getMenu("menu.txt")
        mode.workBook, mode.sells = mode.getLog("sellsLog/")
        mode.initializedScreens = False
        mode.app.setActiveMode(mode.app.logScreen)

########## Different Screens Functions #############  
class Log(Mode):
    def appStarted(mode):
        if(not mode.app.initializeSystem.initializedScreens):
            mode.app.setActiveMode(mode.app.entryScreen)
        else:
            pass

    def keyPressed(mode, event):
        pass

    def timerFired(mode):
        pass

    def redrawAll(mode, canvas):
        Rect, Text = canvas.create_rectangle, canvas.create_text
        pass

class EntryScreen(Mode):
    def appStarted(mode):
        if(not mode.app.initializeSystem.initializedScreens):
            mode.app.setActiveMode(mode.app.newOrderScreen)
        else:
            mode.buttonWidth, mode.buttonHeight = 250, 150
            mode.buttonGap, mode.buttonCx, mode.buttonCy = 750, 350, 650
            mode.buttonFont, mode.calcMssg = "times 28 bold", "Calculadora"
            mode.calcButtonWidth, mode.calcButtonHeight = 250, 75

    def timerFired(mode):
        """Add some sort of interactive background"""
        pass

    def mousePressed(mode, event):
        if mode.app.calculatorOn:
            button = Cmp.calculatorPressed(mode, event.x, event.y)
            if(button == "Exit"):
                mode.app.calculatorOn = False
        else:
            buttonCy, buttonCx, gap = mode.buttonCy, mode.buttonCx, mode.buttonGap
            buttonH, buttonW = mode.buttonHeight, mode.buttonWidth
            buttonHeightRange = range(buttonCy-buttonH, buttonCy+buttonH+1)
            newOrderButton = range(buttonCx-buttonW, buttonCx+buttonW+1)
            ordersLogButton = range(buttonCx-buttonW+gap, buttonCx+buttonW+gap+1)
            if(event.y in buttonHeightRange):
                if(event.x in newOrderButton):
                    mode.app.setActiveMode(mode.app.newOrderScreen)
                elif(event.x in ordersLogButton):
                    mode.app.setActiveMode(mode.app.logScreen)
            elif(event.y < mode.calcButtonHeight and
                event.x > mode.width-mode.calcButtonWidth):
                mode.app.calculatorOn = True

    def createButtons(mode, canvas):
        Rectangle, Text = canvas.create_rectangle, canvas.create_text
        cx, cy = mode.width/2, mode.height/2
        width, height = mode.buttonWidth, mode.buttonHeight
        gap, buttonCx, buttonCy = mode.buttonGap, mode.buttonCx, mode.buttonCy
        for i in range(2):
            if not i%2:
                message = "Nueva\nOrden"
            else:
                message = "Registro\n     de\nOrdenes"
            Rectangle(buttonCx+gap*i-width, buttonCy-height, 
                        buttonCx+gap*i+width, buttonCy+height, fill = "grey")
            Text(buttonCx+gap*i, buttonCy, text = message, font = mode.buttonFont)
            
    def drawCalculatorButton(mode, canvas):
        Rectangle, Text = canvas.create_rectangle, canvas.create_text
        width, height = mode.calcButtonWidth, mode.calcButtonHeight
        Rectangle(mode.width, 0, mode.width-width, height, fill = "white")
        Text(mode.width-width/2, height/2, text = mode.calcMssg, 
                font = "times 30 bold italic")

    def redrawAll(mode, canvas):
        canvas.create_rectangle(0,0, mode.width, mode.height, fill = "black")
        mode.createButtons(canvas)
        if(mode.app.calculatorOn):
            Cmp.drawCalculator(mode, canvas)
        else:
            mode.drawCalculatorButton(canvas)

class NewOrder(Mode):
    def appStarted(mode):
        if(not mode.app.initializeSystem.initializedScreens):
            mode.total = mode.app.currentOrder.getTotal()
            mode.choices = ["Comida", "Bebida", "Postres", "Otros"]
            mode.options = ["Cancelar\n  Orden", "  Ver\nOrden", "Finalizar\n  Orden"]
            mode.buttonHeight, mode.buttonWidth = 75, 250
            mode.buttonCx, mode.buttonCy = 370, 200
            mode.buttonXGap, mode.buttonYGap = 700, 250
            mode.buttonColor, mode.optionButtonsHeight = "light blue", 200
            mode.verifyDecision, mode.verifyCy = False, mode.buttonCy+mode.buttonYGap//2
            mode.verifyWidth, mode.verifyHeight = 250, 150
            mode.verifyButtonCy = mode.verifyCy+(mode.verifyHeight//5*2)
            mode.verifyButtonCx = mode.width//2-mode.verifyWidth//2
            mode.verifyButtonW, mode.verifyButtonH = mode.verifyWidth//4, mode.verifyHeight//4
            mode.order = Order(list())
            mode.screens = [None, mode.app.currentOrderScreen, mode.app.checkoutScreen]
            mode.app.setActiveMode(mode.app.dessertScreen)
        pass

    def pressedButtons(mode, event):
        buttonCy, buttonCx, gap = mode.buttonCy, mode.buttonCx, mode.buttonXGap
        buttonH, buttonW = mode.buttonHeight, mode.buttonWidth
        xGap, yGap = mode.buttonXGap, mode.buttonYGap
        buttonFirstRow = range(buttonCy-buttonH, buttonCy+buttonH+1)
        buttonSecRow = range(buttonCy+yGap-buttonH, buttonCy+yGap+buttonH+1)
        buttonFirstCol = range(buttonCx-buttonW, buttonCx+buttonW+1)
        buttonSecCol = range(buttonCx-buttonW+xGap, buttonCx+buttonW+xGap+1)
        if(event.x in buttonFirstCol):
            if(event.y in buttonFirstRow):
                mode.app.setActiveMode(mode.app.sandwichScreen)
            elif(event.y in buttonSecRow):
                mode.app.setActiveMode(mode.app.dessertsScreen)
        elif(event.x in buttonSecCol):
            if(event.y in buttonFirstRow):
                mode.app.setActiveMode(mode.app.beverageScreen)
            elif(event.y in buttonSecRow):
                mode.app.setActiveMode(mode.app.otherScreen)

    def pressedOptions(mode, event):
        divider = mode.width//3
        screen = mode.screens[event.x//divider]
        if(screen == None):
            Cmp.verifyDecision = True
            return
        else:
            mode.app.setActiveMode(screen)
            # Create a prompt to ask and make sure that
            # the user actually wants to cancel the order
            # and it was not by mistake if so 
            # Call function to clear the current order
            # if not just exit out from the prompt
            
    def mousePressed(mode, event):
        if(Cmp.verifyDecision):
            yRange = range(mode.verifyButtonCy-mode.verifyButtonH, 
                            mode.verifyButtonCy+mode.verifyButtonH+1)
            if(event.y in yRange):
                decision = Cmp.pressedVerifyOption(event)
                if(decision):
                    mode.app.setActiveMode(mode.app.entryScreen)
        else:
            if(event.y >= mode.height-mode.optionButtonsHeight):
                mode.pressedOptions(event)
            else:
                mode.pressedButtons(event)
        
    #### To create transparent rectangles 
    """
    https://newbedev.com/how-to-make-a-tkinter-canvas-rectangle-transparent
    
    """
    def drawOptionButtons(mode, canvas):
        Rect, Oval = canvas.create_rectangle, canvas.create_oval
        Text = canvas.create_text
        buttonWidth, buttonHeight = mode.width//3, mode.optionButtonsHeight
        buttonFont = mode.app.entryScreen.buttonFont
        for i in range(3):
            buttonCx = (buttonWidth//2) + buttonWidth*i
            buttonCy, message = mode.height-(buttonHeight//2), mode.options[i]
            
            if i == 0:
                color = "indian red"
            elif i == 1:
                color = "grey"
            else:
                color = "light green"
                message += "\nTotal: "+str(mode.app.currentOrder.getTotal())+"$"
            Rect(buttonWidth*i, mode.height,  buttonWidth*(i+1), 
                mode.height-buttonHeight, fill = color)
            Text(buttonCx, buttonCy, text = message, font = buttonFont)
      
    def drawButtons(mode, canvas):
        Rect, Oval = canvas.create_rectangle, canvas.create_oval
        Text = canvas.create_text
        buttonW, buttonH = mode.buttonWidth, mode.buttonHeight
        xGap, yGap, color = mode.buttonXGap, mode.buttonYGap, mode.buttonColor
        buttonFont = mode.app.entryScreen.buttonFont
        for i in range(len(mode.choices)):
            message = mode.choices[i]
            buttonCx = mode.buttonCx + xGap*(i%2)
            buttonCy = mode.buttonCy + yGap*(i//2)
            Rect(buttonCx-buttonW, buttonCy-buttonH, 
                buttonCx+buttonW, buttonCy+buttonH, fill = color)
            Text(buttonCx, buttonCy, text = message, font = buttonFont)
        mode.drawOptionButtons(canvas)

    def redrawAll(mode, canvas):
        Rect, Oval = canvas.create_rectangle, canvas.create_oval
        Rect(0, 0, mode.width, mode.height, fill = "black")
        mode.drawButtons(canvas)
        if(Cmp.verifyDecision):
            Cmp.drawVerificationButton(mode, canvas)
    
class Desserts(Mode):
    def appStarted(mode):
        if(not mode.app.initializeSystem.initializedScreens):
            mode.app.setActiveMode(mode.app.otherScreen)
        else:
            pass

    def mousePressed(mode, event):
        pass

    def timerFired(mode):
        pass
        
    def redrawAll(mode, canvas):
        pass
        
    def keyPressed(mode, event):
        pass

class Others(Mode):
    def appStarted(mode):
        if(not mode.app.initializeSystem.initializedScreens):
            mode.app.setActiveMode(mode.app.sandwichScreen)
        else:
            pass

    def mousePressed(mode, event):
        pass

    def timerFired(mode):
        pass
        
    def redrawAll(mode, canvas):
        pass
        
    def keyPressed(mode, event):
        pass

class Sandwiches(Mode):
    def appStarted(mode):
        if(not mode.app.initializeSystem.initializedScreens):
            mode.app.setActiveMode(mode.app.beverageScreen)
        else:
            pass
    
    def timerFired(mode):
        pass

    def keyPressed(mode, event):
        pass

    def redrawAll(mode, canvas):
        pass

class Beverages(Mode):
    def uploadDrinksIcons(mode, path, options):
        mode.iconImages = dict()
        for drink in options:
            newPath = path+drink+"Icon.png"
            mode.iconImages[drink+"Icon"] = mode.loadImage(newPath)
            mode.iconImages[drink+"Icon"] = mode.scaleImage(mode.iconImages[drink+"Icon"], 1.5)
            mode.iconImages[drink+"Icon"].cachedPhotoImage = ImageTk.PhotoImage(mode.iconImages[drink+"Icon"])

    def appStarted(mode):
        if(not mode.app.initializeSystem.initializedScreens):
            mode.cards = mode.app.initializeSystem.getMenuDrinks()
            mode.cardsIcons = mode.uploadDrinksIcons("Images/Icons/", mode.cards)
            mode.app.setActiveMode(mode.app.currentOrderScreen)
            mode.verifyDecision = False
            Cmp.cardCx, Cmp.cardCy = mode.width//4, mode.height//6+40
            Cmp.screenRange = (-200, Cmp.cardCy)
            mode.scrolling = False

    def mousePressed(mode, event):
        if(mode.app.showOptions):
            itemPressed = Cmp.optionsPressed(mode, mode.app.showOptions, event.x, event.y)
            if(itemPressed == "Exit"):
                mode.app.showOptions = False
            elif(itemPressed != None):
                # Need to get price from the calculator for the item and then 
                # create an object for the item and add it to the cart
                mode.app.calculatorOn, mode.app.getPrice = True, True
                mode.calcMssg = itemPressed
                mode.app.showOptions = False
        elif(mode.app.calculatorOn):
            button = Cmp.calculatorPressed(mode, event.x, event.y)
            if(button == "Enter"):
                mode.app.calculatorOn, mode.app.getPrice = False, False
                mode.calcMssg = "Calculadora"
                newDrink = Beverage(mode.calcMssg, int(Cmp.calcNumber))
                mode.app.currentOrder.addItem(newDrink)
            elif(button == "Exit"):
                mode.app.calculatorOn, mode.app.getPrice = False, False
                mode.calcMssg = "Calculadora"
        elif(Cmp.verifyDecision):
            yRange = range(Cmp.verifyButtonCy-Cmp.verifyButtonH, 
                            Cmp.verifyButtonCy+Cmp.verifyButtonH+1)
            if(event.y in yRange):
                decision = Cmp.pressedVerifyOption(event)
                if(decision):
                    mode.app.setActiveMode(mode.app.entryScreen)
        else:
            cardPressed = Cmp.pressedCard(mode, event)

            if(cardPressed != None):
                mode.app.showOptions = mode.cards[cardPressed]
            elif(event.y >= mode.height-mode.app.currentOrderScreen.optionButtonsHeight):
                mode.app.currentOrderScreen.pressedOptions(event)

    def mouseDragged(mode, event):
        if(Cmp.scrolling):
            diff = event.y-Cmp.scrolling
            Cmp.scrollScreen(diff)
            Cmp.scrolling = event.y
        else:
            Cmp.scrolling = event.y

    def mouseReleased(mode, event):
        if(Cmp.scrolling):
            Cmp.scrolling = False

    def redrawAll(mode, canvas):
        Cmp.drawOptionsCards(mode, canvas)
        Cmp.drawIcons(mode, canvas)
        mode.app.currentOrderScreen.drawOptionButtons(canvas)
        if(Cmp.verifyDecision):
            Cmp.drawVerificationButton(mode, canvas)
        if(mode.app.showOptions):
            Cmp.drawOptions(mode, canvas, mode.app.showOptions)
        if(mode.app.calculatorOn):
            Cmp.drawCalculator(mode, canvas)
            

class CurrentOrder(Mode):
    def getNewOrderInfo(mode):
        mode.verifyButtonCy = mode.app.newOrderScreen.verifyButtonCy
        mode.verifyButtonCx = mode.app.newOrderScreen.verifyButtonCx
        mode.verifyButtonW = mode.app.newOrderScreen.verifyButtonW
        mode.buttonWidth = mode.app.newOrderScreen.buttonWidth
        mode.verifyButtonH = mode.app.newOrderScreen.verifyButtonH
        mode.optionButtonsHeight = mode.app.newOrderScreen.optionButtonsHeight

    def appStarted(mode):
        if(not mode.app.initializeSystem.initializedScreens):
            mode.getNewOrderInfo()
            mode.order = mode.app.newOrderScreen.order
            mode.screens = [None, mode.app.newOrderScreen, mode.app.checkoutScreen]
            mode.options = ["Cancelar\n  Orden", "  Volver\n   Atras", "Finalizar\n  Orden"]
            mode.verifyDecision = False
            mode.app.setActiveMode(mode.app.checkoutScreen)
            pass

    def timerFired(mode):
        pass

    def pressedOptions(mode, event):
        divider = mode.width//3
        screen = mode.screens[event.x//divider]
        if(screen == None):
            Cmp.verifyDecision = True
            return
        else:
            mode.app.setActiveMode(screen)

    def mousePressed(mode, event):
        if(Cmp.verifyDecision):
            yRange = range(mode.verifyButtonCy-mode.verifyButtonH, 
                            mode.verifyButtonCy+mode.verifyButtonH+1)
            if(event.y in yRange):
                decision = Cmp.pressedVerifyOption(event)
                if(decision):
                    mode.app.setActiveMode(mode.app.entryScreen)
        else:
            if(event.y >= mode.height-mode.optionButtonsHeight):
                mode.pressedOptions(event)

    def drawOptionButtons(mode, canvas):
        Rect, Oval = canvas.create_rectangle, canvas.create_oval
        Text = canvas.create_text
        buttonWidth, buttonHeight = mode.width//3, mode.optionButtonsHeight
        buttonFont = mode.app.entryScreen.buttonFont
        for i in range(3):
            buttonCx = (buttonWidth//2) + buttonWidth*i
            buttonCy, message = mode.height-(buttonHeight//2), mode.options[i]
            if i == 0:
                color = "indian red"
            elif i == 1:
                color = "grey"
            else:
                color = "light green"
                message += "\nTotal: "+str(mode.app.currentOrder.getTotal())+"$"
            Rect(buttonWidth*i, mode.height,  buttonWidth*(i+1), 
                mode.height-buttonHeight, fill = color)
            Text(buttonCx, buttonCy, text = message, font = buttonFont)
      
    def redrawAll(mode, canvas):
        mode.drawOptionButtons(canvas)
        if(Cmp.verifyDecision):
            Cmp.drawVerificationButton(mode, canvas)
        pass

class Checkout(Mode):
    def appStarted(mode):
        if(not mode.app.initializeSystem.initializedScreens):
            mode.app.initializeSystem.initializedScreens = True
            mode.app.setActiveMode(mode.app.entryScreen)
        else:
            pass

    def timerFired(mode):
        pass

    def keyPressed(mode, event):
        pass

    def redrawAll(mode, canvas):
        pass

########## Menu(Sandwiches/Beverages/Desserts) Screens Functions #############

class Sandwich(object):
    def __init__(self, kind, price):
        self.name = kind
        self.price = price
        pass

class Beverage(object):
    def __init__(self, kind, price):
        self.name = kind
        self.price = price
        pass

class Order(object):
    def __init__(self, cart):
        self.cart = cart
        
    def getTotal(self):
        total = 0
        for item in self.cart:
            total += item.price
        return total

    def getFood(self):
        result = list()
        for order in self.orders:
            if(type(order) == Sandwich):
                result.append(order)
        return result

    def getDrinks(self):
        result = list()
        for order in self.orders:
            if(type(order) == Beverage):
                result.append(order)
        return result
    
    def addItem(self, item):
        self.cart.append(item)

    """
    To get the time for each transaction we do time = now.strftime("%X")
        print("time:", time)
    
    Saving each transaction in a dictionary with the current time as the key.

    with open('data.json', 'w') as fp:
    json.dump(dict, fp, sort_keys=True, indent=4)

    To add Credit card payments:
     https://stories.mlh.io/adding-payments-functionality-to-your-python-app-in-10-minutes-using-the-authorize-net-api-99f5e3e403ab

    """

### Mode superclass has been inherited from cmu_112_graphics 
### http://www.cs.cmu.edu/~112/notes/hw11.html

class MyApp(ModalApp):
    def appStarted(app):
        app.entryScreen = EntryScreen()
        app.logScreen = Log()
        app.newOrderScreen = NewOrder()
        app.sandwichScreen = Sandwiches()
        app.beverageScreen = Beverages()
        app.checkoutScreen = Checkout()
        app.dessertScreen = Desserts()
        app.otherScreen = Others()
        app.currentOrderScreen = CurrentOrder()
        app.initializeSystem = InitializeSystem()
        app.calculatorOn, app.getPrice = False, False
        app.showOptions = False
        app.currentOrder = Order([])
        app.setActiveMode(app.initializeSystem)
        app.timerDelay = 100

app = MyApp(width=1450, height=850)